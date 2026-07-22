#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import os
import zipfile
from pathlib import Path
from typing import Any


ERROR_PATTERNS = [
    "traceback",
    "filenotfounderror",
    "permission denied",
    "no such file",
    "error:",
    "failed",
    "could not",
    "unable to",
    "placeholder",
    "todo",
    "dummy",
    "sample only",
    "not implemented",
]


def load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8", errors="replace"))


def read_success_tsv(path: Path) -> list[dict[str, str]]:
    rows = []
    for line in path.read_text(encoding="utf-8", errors="replace").splitlines():
        if not line.strip():
            continue
        parts = line.split("\t")
        if len(parts) < 5:
            rows.append({
                "task_index": "",
                "task_num": "",
                "attempt": "",
                "duration_sec": "",
                "run_dir": "",
                "parse_problem": f"bad_success_tsv_line: {line}",
            })
            continue

        rows.append({
            "task_index": parts[0],
            "task_num": parts[1],
            "attempt": parts[2],
            "duration_sec": parts[3],
            "run_dir": parts[4],
            "parse_problem": "",
        })
    return rows


def resolve_output_path(raw_path: str, run_dir: Path) -> Path:
    p = Path(raw_path)

    if p.exists():
        return p

    # If status.json stored an absolute path that does not exist in the current view,
    # fall back to run_dir/output/<basename>.
    fallback = run_dir / "output" / p.name
    if fallback.exists():
        return fallback

    # Also try relative path from run_dir/output.
    fallback2 = run_dir / "output" / raw_path
    if fallback2.exists():
        return fallback2

    return p


def inspect_file(path: Path) -> dict[str, Any]:
    result: dict[str, Any] = {
        "path": str(path),
        "exists": path.exists(),
        "size_bytes": None,
        "suffix": path.suffix.lower(),
        "open_ok": False,
        "open_note": "",
        "problems": [],
    }

    if not path.exists():
        result["problems"].append("missing_file")
        return result

    size = path.stat().st_size
    result["size_bytes"] = size

    if size == 0:
        result["problems"].append("zero_byte_file")
        return result

    if size < 512:
        result["problems"].append("very_small_file_under_512_bytes")

    suffix = path.suffix.lower()

    try:
        if suffix == ".xlsx":
            try:
                import openpyxl
                wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
                result["open_ok"] = True
                result["open_note"] = f"sheets={wb.sheetnames}"
                if not wb.sheetnames:
                    result["problems"].append("xlsx_no_sheets")
                wb.close()
            except Exception as e:
                result["problems"].append(f"xlsx_open_failed:{type(e).__name__}:{e}")

        elif suffix == ".docx":
            # docx is a zip package; this is a lightweight structural check.
            if zipfile.is_zipfile(path):
                with zipfile.ZipFile(path) as z:
                    names = set(z.namelist())
                    if "word/document.xml" in names:
                        result["open_ok"] = True
                        result["open_note"] = "valid_docx_zip_with_document_xml"
                    else:
                        result["problems"].append("docx_missing_word_document_xml")
            else:
                result["problems"].append("docx_not_zipfile")

        elif suffix == ".pptx":
            if zipfile.is_zipfile(path):
                with zipfile.ZipFile(path) as z:
                    names = set(z.namelist())
                    slide_count = sum(1 for n in names if n.startswith("ppt/slides/slide") and n.endswith(".xml"))
                    result["open_ok"] = True
                    result["open_note"] = f"valid_pptx_zip_slides={slide_count}"
                    if slide_count == 0:
                        result["problems"].append("pptx_no_slides")
            else:
                result["problems"].append("pptx_not_zipfile")

        elif suffix == ".pdf":
            try:
                import pypdf
                reader = pypdf.PdfReader(str(path))
                n = len(reader.pages)
                result["open_ok"] = True
                result["open_note"] = f"pdf_pages={n}"
                if n == 0:
                    result["problems"].append("pdf_zero_pages")
            except Exception as e:
                result["problems"].append(f"pdf_open_failed:{type(e).__name__}:{e}")

        elif suffix in {".csv", ".tsv"}:
            text = path.read_text(encoding="utf-8", errors="replace")
            result["open_ok"] = True
            lines = text.splitlines()
            result["open_note"] = f"text_lines={len(lines)}"
            if len(lines) <= 1:
                result["problems"].append("csv_or_tsv_has_one_or_fewer_lines")

        elif suffix in {".json", ".jsonl"}:
            if suffix == ".json":
                load_json(path)
                result["open_ok"] = True
                result["open_note"] = "valid_json"
            else:
                good = 0
                bad = 0
                for line in path.read_text(encoding="utf-8", errors="replace").split("\n"):
                    if not line.strip():
                        continue
                    try:
                        json.loads(line)
                        good += 1
                    except Exception:
                        bad += 1
                result["open_ok"] = bad == 0 and good > 0
                result["open_note"] = f"jsonl_good={good}, jsonl_bad={bad}"
                if bad:
                    result["problems"].append("jsonl_bad_lines")

        elif suffix in {".txt", ".md", ".html", ".xml"}:
            text = path.read_text(encoding="utf-8", errors="replace")
            result["open_ok"] = True
            result["open_note"] = f"text_chars={len(text)}"
            if len(text.strip()) < 20:
                result["problems"].append("text_too_short")

        elif suffix in {".png", ".jpg", ".jpeg", ".webp"}:
            try:
                from PIL import Image
                with Image.open(path) as im:
                    result["open_ok"] = True
                    result["open_note"] = f"image_size={im.size}, mode={im.mode}"
                    if im.size[0] < 10 or im.size[1] < 10:
                        result["problems"].append("image_tiny_dimensions")
            except Exception as e:
                result["problems"].append(f"image_open_failed:{type(e).__name__}:{e}")

        elif suffix in {".mp4", ".mov", ".avi", ".wav", ".mp3", ".m4a"}:
            # Structural media validation is harder without ffprobe; at least check non-empty.
            result["open_ok"] = True
            result["open_note"] = "media_file_exists_nonempty_not_deep_validated"

        else:
            result["open_ok"] = True
            result["open_note"] = "unknown_extension_exists_nonempty"

    except Exception as e:
        result["problems"].append(f"unexpected_inspection_error:{type(e).__name__}:{e}")

    return result


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--out-dir", required=True)
    ap.add_argument("--min-output-size", type=int, default=512)
    args = ap.parse_args()

    out_dir = Path(args.out_dir)
    success_tsv = out_dir / "successful_run_dirs.tsv"

    if not success_tsv.exists():
        raise SystemExit(f"[fatal] missing {success_tsv}")

    rows = read_success_tsv(success_tsv)

    audit_rows: list[dict[str, Any]] = []
    file_rows: list[dict[str, Any]] = []

    for row in rows:
        task_num = row["task_num"]
        run_dir = Path(row["run_dir"])

        problems: list[str] = []
        output_files: list[str] = []
        finish_paths: list[str] = []
        status_value = None
        error_value = None
        finish_reason = ""

        if row.get("parse_problem"):
            problems.append(row["parse_problem"])

        status_path = run_dir / "status.json"
        run_metadata_path = run_dir / "run_metadata.json"

        if not run_dir.exists():
            problems.append("run_dir_missing")
        if not status_path.exists():
            problems.append("status_json_missing")
            status = {}
        else:
            try:
                status = load_json(status_path)
                status_value = status.get("status")
                error_value = status.get("error")
                output_files = status.get("output_files") or []
                fp = status.get("finish_params") or {}
                finish_paths = fp.get("paths") or []
                finish_reason = fp.get("reason") or ""
            except Exception as e:
                status = {}
                problems.append(f"status_json_unreadable:{type(e).__name__}:{e}")

        if status_value != "finished":
            problems.append(f"status_not_finished:{status_value}")

        if error_value not in (None, "", "null"):
            problems.append(f"status_error_not_empty:{error_value}")

        if not finish_paths:
            problems.append("finish_paths_empty")

        if not output_files:
            problems.append("output_files_empty")

        lower_finish = finish_reason.lower()
        suspicious_finish_patterns = [
            p for p in ERROR_PATTERNS
            if p in lower_finish and p not in {"error:"}
        ]
        if suspicious_finish_patterns:
            problems.append("suspicious_finish_reason_patterns:" + ",".join(suspicious_finish_patterns))

        # Inspect output files.
        inspected = []
        for raw in output_files:
            p = resolve_output_path(str(raw), run_dir)
            info = inspect_file(p)
            inspected.append(info)

            file_rows.append({
                "task_num": task_num,
                "run_dir": str(run_dir),
                "raw_output_path": str(raw),
                "resolved_output_path": str(p),
                "exists": info["exists"],
                "size_bytes": info["size_bytes"],
                "suffix": info["suffix"],
                "open_ok": info["open_ok"],
                "open_note": info["open_note"],
                "problems": ";".join(info["problems"]),
            })

        if inspected and not all(x["exists"] for x in inspected):
            problems.append("some_output_files_missing")

        if inspected and not all(x["open_ok"] for x in inspected):
            problems.append("some_output_files_failed_open_check")

        tiny = [
            x for x in inspected
            if x["size_bytes"] is not None and x["size_bytes"] < args.min_output_size
        ]
        if tiny:
            problems.append(f"some_output_files_under_{args.min_output_size}_bytes")

        # Lightweight run_metadata check.
        if not run_metadata_path.exists():
            problems.append("run_metadata_missing")
            num_code_exec = None
            num_finish = None
        else:
            try:
                md = load_json(run_metadata_path)
                num_code_exec = len(md.get("code_exec") or [])
                num_finish = len(md.get("finish") or [])
                if num_finish != 1:
                    problems.append(f"finish_call_count_not_1:{num_finish}")
            except Exception as e:
                num_code_exec = None
                num_finish = None
                problems.append(f"run_metadata_unreadable:{type(e).__name__}:{e}")

        audit_rows.append({
            "task_num": task_num,
            "task_index": row.get("task_index"),
            "attempt": row.get("attempt"),
            "duration_sec": row.get("duration_sec"),
            "run_dir": str(run_dir),
            "status": status_value,
            "error": error_value,
            "num_output_files": len(output_files),
            "finish_paths": "|".join(map(str, finish_paths)),
            "output_file_basenames": "|".join(Path(str(x)).name for x in output_files),
            "num_code_exec": num_code_exec,
            "num_finish": num_finish,
            "num_problems": len(problems),
            "problems": ";".join(problems),
        })

    # Write results.
    out_json = out_dir / "successful_outputs_audit.json"
    out_tsv = out_dir / "successful_outputs_audit.tsv"
    out_files_tsv = out_dir / "successful_outputs_file_audit.tsv"

    out_json.write_text(json.dumps(audit_rows, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    with out_tsv.open("w", encoding="utf-8", newline="") as f:
        fields = [
            "task_num",
            "task_index",
            "attempt",
            "duration_sec",
            "status",
            "error",
            "num_output_files",
            "finish_paths",
            "output_file_basenames",
            "num_code_exec",
            "num_finish",
            "num_problems",
            "problems",
            "run_dir",
        ]
        w = csv.DictWriter(f, fieldnames=fields, delimiter="\t")
        w.writeheader()
        for r in audit_rows:
            w.writerow({k: r.get(k, "") for k in fields})

    with out_files_tsv.open("w", encoding="utf-8", newline="") as f:
        fields = [
            "task_num",
            "run_dir",
            "raw_output_path",
            "resolved_output_path",
            "exists",
            "size_bytes",
            "suffix",
            "open_ok",
            "open_note",
            "problems",
        ]
        w = csv.DictWriter(f, fieldnames=fields, delimiter="\t")
        w.writeheader()
        for r in file_rows:
            w.writerow({k: r.get(k, "") for k in fields})

    bad = [r for r in audit_rows if r["num_problems"]]
    good = [r for r in audit_rows if not r["num_problems"]]

    print("=" * 100)
    print("SUCCESSFUL OUTPUT AUDIT")
    print("=" * 100)
    print("out_dir:", out_dir)
    print("successful_task_rows:", len(audit_rows))
    print("clean_by_structural_audit:", len(good))
    print("flagged_by_structural_audit:", len(bad))
    print()
    print("wrote:", out_json)
    print("wrote:", out_tsv)
    print("wrote:", out_files_tsv)

    if bad:
        print()
        print("=" * 100)
        print("FLAGGED TASKS")
        print("=" * 100)
        for r in bad:
            print(f"{r['task_num']}\t{r['problems']}\t{r['run_dir']}")
        raise SystemExit(1)

    print()
    print("[ok] all successful outputs passed structural checks")


if __name__ == "__main__":
    main()
